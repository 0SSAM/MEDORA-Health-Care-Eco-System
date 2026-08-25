from __future__ import annotations

import json
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

API = "https://clinicaltrials.gov/api/v2/studies"
OUT_DIR = Path("/home/ubuntu/bdf-pharma-erp/research_exports")
OUT_DIR.mkdir(parents=True, exist_ok=True)
JSON_PATH = OUT_DIR / "egypt_clinical_trials_raw.json"
XLSX_PATH = OUT_DIR / "egypt_clinical_trials_archive.xlsx"
NOW = datetime.now(timezone.utc).isoformat()


def get_path(obj: dict[str, Any], *path: str, default: Any = "") -> Any:
    cur: Any = obj
    for key in path:
        if not isinstance(cur, dict):
            return default
        cur = cur.get(key)
    return default if cur is None else cur


def as_text(value: Any) -> str:
    if isinstance(value, list):
        return "; ".join(as_text(x) for x in value if x not in (None, ""))
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    return "" if value is None else str(value)


def fetch_all() -> list[dict[str, Any]]:
    studies: list[dict[str, Any]] = []
    token = None
    session = requests.Session()
    while True:
        params = {"query.locn": "Egypt", "pageSize": 100, "format": "json"}
        if token:
            params["pageToken"] = token
        response = session.get(API, params=params, timeout=60)
        response.raise_for_status()
        payload = response.json()
        page = payload.get("studies", [])
        studies.extend(page)
        token = payload.get("nextPageToken")
        if not token:
            break
        time.sleep(0.15)
    return studies


def extract_rows(studies: list[dict[str, Any]]) -> tuple[list[list[str]], list[list[str]]]:
    trial_rows: list[list[str]] = []
    location_rows: list[list[str]] = []
    for study in studies:
        p = study.get("protocolSection", {})
        ident = p.get("identificationModule", {})
        status = p.get("statusModule", {})
        sponsor = p.get("sponsorCollaboratorsModule", {})
        design = p.get("designModule", {})
        conditions = p.get("conditionsModule", {})
        arms = p.get("armsInterventionsModule", {})
        contacts = p.get("contactsLocationsModule", {})
        eligibility = p.get("eligibilityModule", {})
        nct = as_text(ident.get("nctId"))
        locations = contacts.get("locations", []) or []
        egypt_locations = [x for x in locations if as_text(x.get("country")).strip().lower() in {"egypt", "مصر"}]
        trial_rows.append([
            nct,
            as_text(ident.get("briefTitle")),
            as_text(ident.get("officialTitle")),
            as_text(ident.get("acronym")),
            as_text(status.get("overallStatus")),
            as_text(status.get("statusVerifiedDate")),
            as_text(get_path(status, "startDateStruct", "date")),
            as_text(get_path(status, "primaryCompletionDateStruct", "date")),
            as_text(get_path(status, "completionDateStruct", "date")),
            as_text(status.get("studyFirstPostDateStruct", {}).get("date")),
            as_text(status.get("lastUpdatePostDateStruct", {}).get("date")),
            as_text(sponsor.get("leadSponsor", {}).get("name")),
            as_text(sponsor.get("responsibleParty", {}).get("investigatorFullName")),
            as_text(design.get("studyType")),
            as_text(design.get("phases")),
            as_text(design.get("enrollmentInfo", {}).get("count")),
            as_text(conditions.get("conditions")),
            as_text(arms.get("interventions")),
            as_text(eligibility.get("sex")),
            as_text(eligibility.get("minimumAge")),
            as_text(eligibility.get("maximumAge")),
            as_text(eligibility.get("healthyVolunteers")),
            str(len(egypt_locations)),
            f"https://clinicaltrials.gov/study/{nct}" if nct else "",
            NOW,
        ])
        for loc in egypt_locations:
            location_rows.append([
                nct,
                as_text(loc.get("facility")),
                as_text(loc.get("status")),
                as_text(loc.get("city")),
                as_text(loc.get("state")),
                as_text(loc.get("zip")),
                as_text(loc.get("country")),
                as_text(loc.get("contacts")),
                as_text(loc.get("geoPoint")),
                f"https://clinicaltrials.gov/study/{nct}" if nct else "",
                NOW,
            ])
    return trial_rows, location_rows


def write_sheet(ws, headers: list[str], rows: list[list[str]]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col in range(1, ws.max_column + 1):
        width = min(max(12, max((len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, min(ws.max_row, 80) + 1)), default=12) + 2), 42)
        ws.column_dimensions[get_column_letter(col)].width = width


def main() -> None:
    studies = fetch_all()
    JSON_PATH.write_text(json.dumps({"retrievedAt": NOW, "query": "query.locn=Egypt", "studies": studies}, ensure_ascii=False, indent=2), encoding="utf-8")
    trial_rows, location_rows = extract_rows(studies)
    wb = Workbook()
    readme = wb.active
    readme.title = "README"
    readme_rows = [
        ["Dataset", "Egypt-linked clinical-trials archive"],
        ["Retrieved at UTC", NOW],
        ["Primary source", "ClinicalTrials.gov API v2"],
        ["Query", "query.locn=Egypt; pageSize=100; all API pages"],
        ["Study records", len(trial_rows)],
        ["Egypt location records", len(location_rows)],
        ["Coverage statement", "Records returned by the public API query at retrieval time; this is not a guarantee of every Egyptian trial or every national registry."],
        ["Verification rule", "Every row retains NCT identifier, source URL, and retrieval timestamp. Missing source fields remain blank."],
        ["Privacy note", "This archive contains public registry metadata only; it does not contain patient-level data or clinical records."],
    ]
    write_sheet(readme, ["Field", "Value"], readme_rows)
    trial_headers = ["NCT ID", "Brief title", "Official title", "Acronym", "Overall status", "Status verified", "Start date", "Primary completion", "Completion date", "First posted", "Last update posted", "Lead sponsor", "Responsible investigator", "Study type", "Phases", "Enrollment", "Conditions", "Interventions", "Sex", "Minimum age", "Maximum age", "Healthy volunteers", "Egypt location count", "Source URL", "Retrieved at UTC"]
    loc_headers = ["NCT ID", "Facility", "Location status", "City", "State", "ZIP", "Country", "Contacts", "Geo point", "Source URL", "Retrieved at UTC"]
    write_sheet(wb.create_sheet("Trials"), trial_headers, trial_rows)
    write_sheet(wb.create_sheet("Egypt Locations"), loc_headers, location_rows)
    sources = wb.create_sheet("Sources")
    write_sheet(sources, ["Source", "URL", "Role", "Access note"], [
        ["ClinicalTrials.gov API v2", API, "Primary public registry data", "Queried with query.locn=Egypt; public registry metadata only."],
        ["WHO ICTRP", "https://www.who.int/tools/clinical-trials-registry-platform", "Registry discovery and methodology", "WHO states ICTRP is not itself a registry and links primary registries."],
        ["WHO ICTRP search portal", "https://trialsearch.who.int/", "Cross-registry discovery", "Use as a complementary source; coverage may differ from ClinicalTrials.gov."],
    ])
    wb.save(XLSX_PATH)
    print(json.dumps({"studies": len(trial_rows), "egypt_locations": len(location_rows), "xlsx": str(XLSX_PATH), "raw_json": str(JSON_PATH)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
