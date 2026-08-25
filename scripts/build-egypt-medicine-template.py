from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from pathlib import Path

out = Path('/home/ubuntu/bdf-pharma-erp/docs/data/egypt-medicine-register-source-safe-template.xlsx')
out.parent.mkdir(parents=True, exist_ok=True)
wb = Workbook()
ws = wb.active
ws.title = 'Medicine Register'
headers = ['record_id','name_ar','name_en','active_ingredient','strength','dosage_form','manufacturer','registration_number','registration_status','source_url','source_published_at','verified_at','verification_method','jurisdiction','notes']
ws.append(headers)
for cell in ws[1]:
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor='0B7285')
ws.freeze_panes = 'A2'
ws.auto_filter.ref = 'A1:O1'
for column in ws.columns:
    ws.column_dimensions[column[0].column_letter].width = min(max(len(str(column[0].value or '')) + 2, 14), 28)

method = wb.create_sheet('Methodology')
method.append(['Field','Value'])
method.append(['Status','Template only — no medicine records included'])
method.append(['Jurisdiction','Egypt'])
method.append(['Primary-source requirement','Each row must cite a reproducible official source URL'])
method.append(['Verification requirement','Record verification timestamp and method before use in regulated workflows'])
method.append(['Current limitation','A reproducible official EDA bulk register was not available during this build'])
method.append(['Safety rule','Do not populate from unverified commercial lists or inferred names'])
method.append(['Clinical trials archive','Intentionally skipped per user instruction'])
for cell in method[1]:
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor='0B7285')
method.column_dimensions['A'].width = 32
method.column_dimensions['B'].width = 100

sources = wb.create_sheet('Source Register')
sources.append(['source_url','authority','scope','accessed_at','status','limitations'])
sources.append(['https://edaegypt.gov.eg/en/the-regulatory-reference-of-the-egyptian-drug-authority-eda/regulatory-guidelines/ca-of-pharmaceutical-products/','Egyptian Drug Authority (EDA)','Regulatory guidelines for pharmaceutical products','2026-08-14','Reviewed','Guidelines page; not a complete machine-readable medicine register'])
for cell in sources[1]:
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor='0B7285')
for col, width in {'A':88,'B':30,'C':42,'D':16,'E':16,'F':60}.items():
    sources.column_dimensions[col].width = width

wb.save(out)
print(out)
